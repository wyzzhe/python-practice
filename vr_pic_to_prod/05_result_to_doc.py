#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
整理分析结果，生成商品分析和店铺分析文档
"""

import os
import json
import argparse
import re
import shutil
from typing import List, Dict, Optional
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# ==================== 配置项 ====================
DEFAULT_ANALYSIS_DIR = os.path.join(
    os.path.dirname(__file__),
    'analysis_results'
)

DEFAULT_RECOMMENDED_FILE = os.path.join(
    os.path.dirname(__file__),
    'recommended_views/recommended_products.json'
)

DEFAULT_OUTPUT_DIR = os.path.join(
    os.path.dirname(__file__),
    'output_docs'
)

DEFAULT_BRAND_LIST_FILE = os.path.join(
    os.path.dirname(__file__),
    'brand_list.py'
)


def load_brand_categories(brand_list_file: str = None) -> Dict[str, str]:
    """
    从 brand_list.py 解析品牌分类信息
    
    Returns:
        {brand_name: category_name}
    """
    if brand_list_file is None:
        brand_list_file = DEFAULT_BRAND_LIST_FILE
    
    brand_category_map = {}
    
    if not os.path.exists(brand_list_file):
        print(f'[WARN] 品牌列表文件不存在: {brand_list_file}')
        return brand_category_map
    
    try:
        with open(brand_list_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 使用正则表达式解析分类和品牌
        # 匹配格式: # 分类类\n    "品牌1", "品牌2",
        pattern = r'#\s*([^\n#]+类)\s*\n\s*((?:"[^"]+",?\s*)+)'
        matches = re.findall(pattern, content, re.MULTILINE)
        
        for category, brands_str in matches:
            category = category.strip()
            # 提取品牌名称（去除引号和逗号）
            brand_names = re.findall(r'"([^"]+)"', brands_str)
            for brand in brand_names:
                brand_category_map[brand] = category
        
        print(f'[INFO] 已加载 {len(brand_category_map)} 个品牌的分类信息')
    except Exception as ex:
        print(f'[ERROR] 解析品牌列表文件失败: {ex}')
    
    return brand_category_map


def auto_categorize_brand(brand_name: str) -> str:
    """
    根据品牌名称自动推断分类
    
    Args:
        brand_name: 品牌名称
    
    Returns:
        分类名称
    """
    brand_upper = brand_name.upper()
    
    # 已知品牌分类映射（扩展列表）
    known_brands = {
        # 美妆护肤类
        'LAMER': '美妆护肤类',
        'SK-2': '美妆护肤类',
        'SKINCEUTICALS': '美妆护肤类',
        'SKIN CEUTICALS': '美妆护肤类',
        'CLEDEPEAU': '美妆护肤类',
        'CLARINS': '美妆护肤类',
        'DECORTE': '美妆护肤类',
        'DIOR': '美妆护肤类',
        'GUERLAIN': '美妆护肤类',
        'VALMONT': '美妆护肤类',
        'SISLY': '美妆护肤类',
        'LAPRAIRIE': '美妆护肤类',
        'WHOO': '美妆护肤类',
        'HIEIDO': '美妆护肤类',
        'TOMFORD': '美妆护肤类',
        'TOM FORD': '美妆护肤类',
        'HOURGLASS': '美妆护肤类',
        'BOBBIBROWN': '美妆护肤类',
        'BOBBI BROWN': '美妆护肤类',
        'ACQUADIPARMA': '美妆护肤类',
        'ACQUA DI PARMA': '美妆护肤类',
        'JOMALONE': '美妆护肤类',
        "L'OCCITANE": '美妆护肤类',
        'GIVENCHY': '美妆护肤类',
        '雅诗兰黛': '美妆护肤类',
        
        # 珠宝首饰类
        'BVLGARI': '珠宝首饰类',
        'TUDOR': '珠宝首饰类',
        'OMEGA': '珠宝首饰类',
        'IWC': '珠宝首饰类',
        'CAURTIER': '珠宝首饰类',
        'CARTIER': '珠宝首饰类',
        'CUCCI': '珠宝首饰类',
        'GUCCI': '珠宝首饰类',
        'APM': '珠宝首饰类',
        '老庙': '珠宝首饰类',
        '老凤祥': '珠宝首饰类',
        '诗普琳': '珠宝首饰类',
        '中国珠宝': '珠宝首饰类',
        '周生生': '珠宝首饰类',
        '潮宏基': '珠宝首饰类',
        '周大生': '珠宝首饰类',
        'DR': '珠宝首饰类',
        '中国黄金': '珠宝首饰类',
        '珠利莱': '珠宝首饰类',
        
        # 时尚服饰类
        'ARMANI': '时尚服饰类',
        'MO&CO': '时尚服饰类',
        'MO AND CO': '时尚服饰类',
        '太平鸟': '时尚服饰类',
        '13DEMARZO': '时尚服饰类',
        '13DE MARZO': '时尚服饰类',
        'WE11DONE': '时尚服饰类',
        'ELAND': '时尚服饰类',
        'E.LAND': '时尚服饰类',
        
        # 箱包类
        'DISSONA': '箱包类',
        "ST&SAT": '箱包类',
        'ST AND SAT': '箱包类',
        'GG-CC': '箱包类',
        'GG CC': '箱包类',
        
        # 鞋履类
        # BIRKENSTOCK 已在 brand_list.py 中
        
        # 运动休闲类
        'JORDAN': '运动休闲类',
        'AIR JORDAN': '运动休闲类',
        '李宁': '运动休闲类',
        
        # 餐饮类
        '兰湘子': '餐饮类',
        
        # 汽车类
        '蔚来': '汽车类',
        '智己': '汽车类',
        
        # 家居用品类
        # 睿锦尚品 已在 brand_list.py 中
        
        # 儿童用品类
        '阿吉豆': '儿童用品类',
        
        # 其他已知品牌
        '浮光之秋': '时尚服饰类',
    }
    
    # 首先检查完全匹配
    if brand_name in known_brands:
        return known_brands[brand_name]
    if brand_upper in known_brands:
        return known_brands[brand_upper]
    
    # 关键字匹配
    keywords = {
        '美妆护肤类': ['SK', 'LAMER', 'CLARINS', 'DIOR', 'LANCOME', 'HERMES', 'WHOO', 'VALMONT', 
                      'LA PRAIRIE', 'TOM FORD', 'HOURGLASS', 'BOBBI', 'ACQUA', 'JOMALONE', 
                      'OCCITANE', 'GIVENCHY', 'GUERLAIN', '雅诗兰黛', 'CPB', 'CLEDEPEAU', 'DECORTE'],
        '珠宝首饰类': ['珠宝', '黄金', '钻石', '周', '老', 'DR', 'BVLGARI', 'TIFFANY', 'CARTIER', 
                      'OMEGA', 'IWC', 'TUDOR', 'APM', 'HEFANG', '珠利', '诗普'],
        '时尚服饰类': ['太平鸟', 'MO', 'ARMANI', '13DE', 'MARZO', 'WE11', 'ELAND', '浮光'],
        '箱包类': ['MCM', 'BALLY', 'BOOS', 'DISSONA', 'ST', 'GG', 'CC', 'TUMI', 'COACH'],
        '鞋履类': ['BIRKENSTOCK', 'BIRKEN', 'BOSS'],
        '运动休闲类': ['JORDAN', '耐克', 'NIKE', '安踏', 'ANTA', '李宁', 'LI NING', 'ADIDAS'],
        '餐饮类': ['EAT', 'O\'EAT', '兰湘', '星巴克', 'STARBUCKS', '咖啡', 'COFFEE'],
        '汽车类': ['蔚来', '特斯拉', 'TESLA', '智己', 'BYD', '理想'],
        '家居用品类': ['睿锦', '宜家', 'IKEA'],
        '自行车类': ['BROMPTON', 'SPECIALIZED'],
    }
    
    for category, key_list in keywords.items():
        for key in key_list:
            if key in brand_upper or key in brand_name:
                return category
    
    # 如果都没有匹配到，返回"其他类"
    return '其他类'


def get_brand_category(brand_name: str, brand_category_map: Dict[str, str]) -> str:
    """
    获取品牌分类，如果未找到则自动分类
    
    Args:
        brand_name: 品牌名称
        brand_category_map: 从 brand_list.py 加载的分类映射
    
    Returns:
        分类名称
    """
    # 首先从映射中查找
    category = brand_category_map.get(brand_name)
    if category:
        return category
    
    # 如果未找到，使用自动分类
    category = auto_categorize_brand(brand_name)
    print(f'[INFO] 品牌 "{brand_name}" 未在 brand_list.py 中找到，自动分类为: {category}')
    return category


def load_recommended_products(recommended_file: str) -> Dict[str, List[Dict]]:
    """
    加载推荐商品信息，按 (brand, panorama_id, name) 建立索引
    
    Returns:
        {(brand, panorama_id, name): vr_link}
    """
    vr_links = {}
    
    if not os.path.exists(recommended_file):
        print(f'[WARN] 推荐商品文件不存在: {recommended_file}')
        return vr_links
    
    try:
        with open(recommended_file, 'r', encoding='utf-8') as f:
            products = json.load(f)
        
        for product in products:
            key = (
                product.get('brand', ''),
                product.get('panorama_id', 0),
                product.get('name', '')
            )
            vr_links[key] = product.get('vr_link', '')
        
        print(f'[INFO] 已加载 {len(vr_links)} 个推荐商品的 VR 链接')
    except Exception as e:
        print(f'[ERROR] 加载推荐商品文件失败: {e}')
    
    return vr_links


def load_analysis_results(analysis_dir: str) -> List[Dict]:
    """加载所有分析结果"""
    results = []
    
    if not os.path.isdir(analysis_dir):
        print(f'[ERROR] 目录不存在: {analysis_dir}')
        return results
    
    for filename in os.listdir(analysis_dir):
        if not filename.endswith('_analysis.json'):
            continue
        
        file_path = os.path.join(analysis_dir, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                results.append(data)
        except Exception as e:
            print(f'[ERROR] 读取文件失败 {filename}: {e}')
    
    return results


def copy_image_to_output(image_path: str, output_dir: str) -> Optional[str]:
    """
    将图片复制到输出目录的 images 文件夹中
    
    Args:
        image_path: 原始图片路径
        output_dir: 输出目录
    
    Returns:
        新的相对路径（相对于输出目录），如果复制失败返回 None
    """
    if not os.path.exists(image_path):
        return None
    
    # 创建 images 子目录
    images_dir = os.path.join(output_dir, 'images')
    os.makedirs(images_dir, exist_ok=True)
    
    # 获取文件名
    filename = os.path.basename(image_path)
    
    # 如果文件名已经存在，添加序号避免冲突
    dest_path = os.path.join(images_dir, filename)
    counter = 1
    base_name, ext = os.path.splitext(filename)
    while os.path.exists(dest_path):
        new_filename = f"{base_name}_{counter}{ext}"
        dest_path = os.path.join(images_dir, new_filename)
        counter += 1
    
    try:
        shutil.copy2(image_path, dest_path)
        # 返回相对路径（相对于输出目录）
        rel_path = os.path.relpath(dest_path, output_dir)
        return rel_path.replace('\\', '/')  # 统一使用 / 分隔符
    except Exception as e:
        print(f'[WARN] 复制图片失败 {image_path}: {e}')
        return None


def get_image_display_path(image_path: str, relative_to: str = None, output_dir: str = None) -> str:
    """
    转换图片路径，如果指定了输出目录，则复制图片并返回相对路径；否则返回原路径或相对路径
    
    Args:
        image_path: 图片绝对路径
        relative_to: 相对路径的基准目录（不使用输出目录时）
        output_dir: 输出目录（如果提供，会复制图片）
    """
    if output_dir:
        # 复制图片到输出目录并返回相对路径
        rel_path = copy_image_to_output(image_path, output_dir)
        if rel_path:
            return rel_path
    
    # 如果没有输出目录或复制失败，使用原来的逻辑
    if os.path.exists(image_path):
        if relative_to:
            try:
                rel_path = os.path.relpath(image_path, relative_to)
                return rel_path.replace('\\', '/')  # 统一使用 / 分隔符
            except Exception:
                pass
        return image_path.replace('\\', '/')
    return image_path.replace('\\', '/')


def format_product_info(product: Dict) -> str:
    """格式化商品信息为 Markdown（用于表格单元格）"""
    lines = []
    
    # 商品名称（推荐商品加标记）
    name = product.get('name', '')
    is_recommended = product.get('is_recommended', False)
    if is_recommended:
        name = f"**{name}** ⭐ (推荐)"
    lines.append(f"• **名称**: {name}")
    
    # 其他信息
    if product.get('type'):
        lines.append(f"• **类型**: {product['type']}")
    
    if product.get('colors'):
        colors = '、'.join(product['colors'])
        lines.append(f"• **颜色**: {colors}")
    
    if product.get('materials'):
        materials = '、'.join(product['materials'])
        lines.append(f"• **材质**: {materials}")
    
    if product.get('location'):
        lines.append(f"• **位置**: {product['location']}")
    
    if product.get('view_direction'):
        direction_map = {'f': '前', 'b': '后', 'l': '左', 'r': '右'}
        direction = direction_map.get(product['view_direction'], product['view_direction'])
        lines.append(f"• **视图方向**: {direction}")
    
    if product.get('position_3d'):
        pos = product['position_3d']
        if pos and isinstance(pos, dict):
            lines.append(f"• **3D坐标**: ({pos.get('x', 0)}, {pos.get('y', 0)}, {pos.get('z', 0)})")
    
    # 使用 <br> 分隔，适合表格显示
    return '<br>'.join(lines)


def generate_product_section(results: List[Dict], vr_links: Dict, brand_category_map: Dict[str, str], output_dir: str = None) -> str:
    """生成商品分析部分"""
    lines = []
    lines.append('# 第一部分：商品分析\n')
    lines.append('本部分按店铺和点位展示所有检测到的商品信息。推荐商品已标记 ⭐ 并包含 VR 链接。\n')
    
    # 按分类分组品牌
    category_brands = defaultdict(list)
    for brand_result in results:
        brand_name = brand_result.get('brand', '')
        category = get_brand_category(brand_name, brand_category_map)
        category_brands[category].append(brand_result)
    
    # 按分类遍历
    for category in sorted(category_brands.keys()):
        lines.append(f'\n## {category}\n')
        
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        for brand_result in brand_results:
            brand_name = brand_result.get('brand', '')
            product_results = brand_result.get('product_results', [])
            
            if not product_results:
                continue
            
            lines.append(f'\n### {brand_name}\n')
            
            # 品牌级别只输出一次表头
            lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 商品信息 | VR链接 |')
            lines.append('|---------|---------|---------|---------|---------|---------|')
            
            for point_result in product_results:
                panorama_id = point_result.get('panorama_id', '')
                images = point_result.get('images', [])
                products = point_result.get('products', [])
                
                if not products:
                    continue
                
                # 准备图片链接（四个视角）
                image_links = []
                for img_path in sorted(images):
                    rel_path = get_image_display_path(img_path, output_dir=output_dir)
                    # 判断方向
                    direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                    direction = '未知'
                    for key, value in direction_map.items():
                        if img_path.endswith(key):
                            direction = value
                            break
                    image_links.append(f"[{direction}]({rel_path})")
                
                images_cell = '<br>'.join(image_links) if image_links else '无图片'
                
                # 准备商品信息
                products_info = []
                vr_links_cell = []
                for product in products:
                    product_text = format_product_info(product)
                    products_info.append(product_text)
                    
                    # 获取 VR 链接
                    vr_link = None
                    if product.get('is_recommended', False):
                        key = (brand_name, panorama_id, product.get('name', ''))
                        vr_link = vr_links.get(key)
                    
                    if vr_link:
                        vr_links_cell.append(f"[{product.get('name', '')}]({vr_link})")
                    elif product.get('is_recommended', False):
                        vr_links_cell.append(f"{product.get('name', '')} (未找到)")
                
                products_cell = '<br><br>'.join(products_info) if products_info else '无商品'
                vr_links_cell_text = '<br>'.join(vr_links_cell) if vr_links_cell else '-'
                
                # 输出表格行（添加分类）
                category = get_brand_category(brand_name, brand_category_map)
                lines.append(f'| {category} | {brand_name} | {panorama_id} | {images_cell} | {products_cell} | {vr_links_cell_text} |')
    
    return '\n'.join(lines)


def generate_summary_product_section(results: List[Dict], vr_links: Dict, brand_category_map: Dict[str, str], output_dir: str = None) -> str:
    """生成精简版商品分析部分（每个类别选一个店铺，每个店铺选5个点，每个点选3个推荐商品和2个不推荐商品）"""
    lines = []
    lines.append('# 第一部分：商品分析（精简版）\n')
    lines.append('本部分展示精简版本：每个类别选择一个店铺，每个店铺选择5个点位，每个点位选择3个推荐商品和2个不推荐商品。推荐商品已标记 ⭐ 并包含 VR 链接。\n')
    
    # 按分类分组品牌
    category_brands = defaultdict(list)
    for brand_result in results:
        brand_name = brand_result.get('brand', '')
        category = get_brand_category(brand_name, brand_category_map)
        category_brands[category].append(brand_result)
    
    # 按分类遍历，每个分类只选第一个品牌
    for category in sorted(category_brands.keys()):
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        if not brand_results:
            continue
        
        # 每个分类选择第一个品牌
        brand_result = brand_results[0]
        brand_name = brand_result.get('brand', '')
        product_results = brand_result.get('product_results', [])
        
        if not product_results:
            continue
        
        lines.append(f'\n## {category} - {brand_name}\n')
        
        # 品牌级别只输出一次表头
        lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 商品信息 | VR链接 |')
        lines.append('|---------|---------|---------|---------|---------|---------|')
        
        # 选择最多5个点位
        selected_points = product_results[:5]
        
        for point_result in selected_points:
            panorama_id = point_result.get('panorama_id', '')
            images = point_result.get('images', [])
            products = point_result.get('products', [])
            
            if not products:
                continue
            
            # 准备图片链接（四个视角）
            image_links = []
            for img_path in sorted(images):
                rel_path = get_image_display_path(img_path, output_dir=output_dir)
                # 判断方向
                direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                direction = '未知'
                for key, value in direction_map.items():
                    if img_path.endswith(key):
                        direction = value
                        break
                image_links.append(f"[{direction}]({rel_path})")
            
            images_cell = '<br>'.join(image_links) if image_links else '无图片'
            
            # 准备商品信息：每个点位选3个推荐商品和2个不推荐商品
            recommended_products = [p for p in products if p.get('is_recommended', False)]
            non_recommended_products = [p for p in products if not p.get('is_recommended', False)]
            
            # 选择3个推荐商品和2个不推荐商品
            selected_recommended = recommended_products[:3]
            selected_non_recommended = non_recommended_products[:2]
            selected_products = selected_recommended + selected_non_recommended
            
            products_info = []
            vr_links_cell = []
            for product in selected_products:
                product_text = format_product_info(product)
                products_info.append(product_text)
                
                # 获取 VR 链接
                vr_link = None
                if product.get('is_recommended', False):
                    key = (brand_name, panorama_id, product.get('name', ''))
                    vr_link = vr_links.get(key)
                
                if vr_link:
                    vr_links_cell.append(f"[{product.get('name', '')}]({vr_link})")
                elif product.get('is_recommended', False):
                    vr_links_cell.append(f"{product.get('name', '')} (未找到)")
            
            products_cell = '<br><br>'.join(products_info) if products_info else '无商品'
            vr_links_cell_text = '<br>'.join(vr_links_cell) if vr_links_cell else '-'
            
            # 输出表格行（添加分类）
            category = get_brand_category(brand_name, brand_category_map)
            lines.append(f'| {category} | {brand_name} | {panorama_id} | {images_cell} | {products_cell} | {vr_links_cell_text} |')
    
    return '\n'.join(lines)


def generate_summary_store_section(results: List[Dict], brand_category_map: Dict[str, str], output_dir: str = None) -> str:
    """生成精简版店铺分析部分（每个类别选一个店铺，每个店铺选5个点）"""
    lines = []
    lines.append('\n\n# 第二部分：店铺分析（精简版）\n')
    lines.append('本部分展示精简版本：每个类别选择一个店铺，每个店铺选择5个点位。\n')
    
    # 按分类分组品牌
    category_brands = defaultdict(list)
    for brand_result in results:
        brand_name = brand_result.get('brand', '')
        category = get_brand_category(brand_name, brand_category_map)
        category_brands[category].append(brand_result)
    
    # 按分类遍历，每个分类只选第一个品牌
    for category in sorted(category_brands.keys()):
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        if not brand_results:
            continue
        
        # 每个分类选择第一个品牌
        brand_result = brand_results[0]
        brand_name = brand_result.get('brand', '')
        store_analysis = brand_result.get('store_analysis', {})
        product_results = brand_result.get('product_results', [])
        
        if not store_analysis:
            continue
        
        lines.append(f'\n## {category} - {brand_name}\n')
        
        # 收集所有点位图片
        all_point_images = {}
        for point_result in product_results:
            panorama_id = point_result.get('panorama_id', '')
            images = point_result.get('images', [])
            if images:
                all_point_images[panorama_id] = {
                    'images': images
                }
        
        # 选择最多5个点位
        selected_panorama_ids = sorted(all_point_images.keys())[:5]
        
        # 为每个有图片的点位生成一行
        if selected_panorama_ids:
            lines.append('\n#### 店铺点位信息\n')
            lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 店铺信息 |')
            lines.append('|---------|---------|---------|---------|---------|')
            
            for i, panorama_id in enumerate(selected_panorama_ids):
                point_data = all_point_images[panorama_id]
                images = point_data['images']
                
                # 准备图片链接
                image_links = []
                for img_path in sorted(images):
                    rel_path = get_image_display_path(img_path)
                    direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                    direction = '未知'
                    for key, value in direction_map.items():
                        if img_path.endswith(key):
                            direction = value
                            break
                    image_links.append(f"[{direction}]({rel_path})")
                
                images_cell = '<br>'.join(image_links) if image_links else '无图片'
                
                # 准备店铺信息（第一行显示完整信息，其他行显示 "-"）
                store_info_lines = []
                if i == 0:
                    # 店铺基本信息
                    if store_analysis.get('store_category'):
                        store_info_lines.append(f"**店铺类别**: {store_analysis['store_category']}")
                    
                    if store_analysis.get('price_positioning'):
                        store_info_lines.append(f"**价格定位**: {store_analysis['price_positioning']}")
                    
                    if store_analysis.get('target_customers'):
                        customers = '、'.join(store_analysis['target_customers'])
                        store_info_lines.append(f"**目标客户**: {customers}")
                    
                    # 店铺环境
                    store_env = store_analysis.get('store_env', {})
                    if store_env:
                        if store_env.get('style'):
                            styles = '、'.join(store_env['style']) if isinstance(store_env['style'], list) else store_env['style']
                            store_info_lines.append(f"**风格**: {styles}")
                        
                        if store_env.get('lighting'):
                            store_info_lines.append(f"**照明**: {store_env['lighting']}")
                        
                        if store_env.get('spatial_layout'):
                            store_info_lines.append(f"**空间布局**: {store_env['spatial_layout']}")
                        
                        if store_env.get('overall_feeling'):
                            store_info_lines.append(f"**整体感觉**: {store_env['overall_feeling']}")
                        
                        if store_env.get('display_method'):
                            methods = store_env['display_method']
                            if isinstance(methods, list):
                                methods_text = '<br>'.join([f"  - {m}" for m in methods])
                            else:
                                methods_text = str(methods)
                            store_info_lines.append(f"**陈列方式**:<br>{methods_text}")
                    
                    # 购物体验
                    shopping_exp = store_analysis.get('store_env', {}).get('shopping_experience', {})
                    if shopping_exp:
                        if shopping_exp.get('has_try_on_area') is not None:
                            store_info_lines.append(f"**有试穿区**: {'是' if shopping_exp['has_try_on_area'] else '否'}")
                        
                        if shopping_exp.get('has_photo_spots') is not None:
                            store_info_lines.append(f"**有拍照点**: {'是' if shopping_exp['has_photo_spots'] else '否'}")
                
                store_info_cell = '<br>'.join(store_info_lines) if store_info_lines else '-'
                
                # 输出表格行（添加分类）
                category = get_brand_category(brand_name, brand_category_map)
                lines.append(f'| {category} | {brand_name} | {panorama_id} | {images_cell} | {store_info_cell} |')
        else:
            # 如果没有点位图片，只显示店铺信息
            lines.append('\n#### 店铺整体信息\n')
            lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 店铺信息 |')
            lines.append('|---------|---------|---------|---------|---------|')
            
            store_info_lines = []
            if store_analysis.get('store_category'):
                store_info_lines.append(f"**店铺类别**: {store_analysis['store_category']}")
            
            store_info_cell = '<br>'.join(store_info_lines) if store_info_lines else '无信息'
            category = get_brand_category(brand_name, brand_category_map)
            lines.append(f'| {category} | {brand_name} | - | - | {store_info_cell} |')
    
    return '\n'.join(lines)


def generate_store_section(results: List[Dict], brand_category_map: Dict[str, str], output_dir: str = None) -> str:
    """生成店铺分析部分"""
    lines = []
    lines.append('\n\n# 第二部分：店铺分析\n')
    lines.append('本部分展示各店铺的整体环境分析信息。\n')
    
    # 按分类分组品牌
    category_brands = defaultdict(list)
    for brand_result in results:
        brand_name = brand_result.get('brand', '')
        category = get_brand_category(brand_name, brand_category_map)
        category_brands[category].append(brand_result)
    
    # 按分类遍历
    for category in sorted(category_brands.keys()):
        lines.append(f'\n## {category}\n')
        
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        for brand_result in brand_results:
            brand_name = brand_result.get('brand', '')
            store_analysis = brand_result.get('store_analysis', {})
            product_results = brand_result.get('product_results', [])
            
            if not store_analysis:
                continue
            
            lines.append(f'\n### {brand_name}\n')
            
            # 收集所有点位图片
            all_point_images = {}
            for point_result in product_results:
                panorama_id = point_result.get('panorama_id', '')
                seq_id = point_result.get('seq_id', '')
                images = point_result.get('images', [])
                if images:
                    all_point_images[panorama_id] = {
                        'seq_id': seq_id,
                        'images': images
                    }
            
            # 为每个有图片的点位生成一行
            if all_point_images:
                lines.append('\n#### 店铺点位信息\n')
                lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 店铺信息 |')
                lines.append('|---------|---------|---------|---------|---------|')
            
            for panorama_id in sorted(all_point_images.keys()):
                point_data = all_point_images[panorama_id]
                images = point_data['images']
                
                # 准备图片链接
                image_links = []
                for img_path in sorted(images):
                    rel_path = get_image_display_path(img_path)
                    direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                    direction = '未知'
                    for key, value in direction_map.items():
                        if img_path.endswith(key):
                            direction = value
                            break
                    image_links.append(f"[{direction}]({rel_path})")
                
                images_cell = '<br>'.join(image_links) if image_links else '无图片'
                
                # 准备店铺信息（第一行显示完整信息，其他行显示 "-"）
                store_info_lines = []
                if panorama_id == sorted(all_point_images.keys())[0]:
                    # 店铺基本信息
                    if store_analysis.get('store_category'):
                        store_info_lines.append(f"**店铺类别**: {store_analysis['store_category']}")
                    
                    if store_analysis.get('price_positioning'):
                        store_info_lines.append(f"**价格定位**: {store_analysis['price_positioning']}")
                    
                    if store_analysis.get('target_customers'):
                        customers = '、'.join(store_analysis['target_customers'])
                        store_info_lines.append(f"**目标客户**: {customers}")
                    
                    # 店铺环境
                    store_env = store_analysis.get('store_env', {})
                    if store_env:
                        if store_env.get('style'):
                            styles = '、'.join(store_env['style']) if isinstance(store_env['style'], list) else store_env['style']
                            store_info_lines.append(f"**风格**: {styles}")
                        
                        if store_env.get('lighting'):
                            store_info_lines.append(f"**照明**: {store_env['lighting']}")
                        
                        if store_env.get('spatial_layout'):
                            store_info_lines.append(f"**空间布局**: {store_env['spatial_layout']}")
                        
                        if store_env.get('overall_feeling'):
                            store_info_lines.append(f"**整体感觉**: {store_env['overall_feeling']}")
                        
                        if store_env.get('display_method'):
                            methods = store_env['display_method']
                            if isinstance(methods, list):
                                methods_text = '<br>'.join([f"  - {m}" for m in methods])
                            else:
                                methods_text = str(methods)
                            store_info_lines.append(f"**陈列方式**:<br>{methods_text}")
                    
                    # 购物体验
                    shopping_exp = store_analysis.get('store_env', {}).get('shopping_experience', {})
                    if shopping_exp:
                        if shopping_exp.get('has_try_on_area') is not None:
                            store_info_lines.append(f"**有试穿区**: {'是' if shopping_exp['has_try_on_area'] else '否'}")
                        
                        if shopping_exp.get('has_photo_spots') is not None:
                            store_info_lines.append(f"**有拍照点**: {'是' if shopping_exp['has_photo_spots'] else '否'}")
                
                store_info_cell = '<br>'.join(store_info_lines) if store_info_lines else '-'
                
                # 输出表格行（添加分类）
                category = get_brand_category(brand_name, brand_category_map)
                lines.append(f'| {category} | {brand_name} | {panorama_id} | {images_cell} | {store_info_cell} |')
        else:
            # 如果没有点位图片，只显示店铺信息
            lines.append('\n#### 店铺整体信息\n')
            lines.append('| 店铺分类 | 店铺名称 | 店铺点位 | 点位图片 | 店铺信息 |')
            lines.append('|---------|---------|---------|---------|---------|')
            
            store_info_lines = []
            if store_analysis.get('store_category'):
                store_info_lines.append(f"**店铺类别**: {store_analysis['store_category']}")
            # ... 其他信息（同上）
            
            store_info_cell = '<br>'.join(store_info_lines) if store_info_lines else '无信息'
            category = get_brand_category(brand_name, brand_category_map)
            lines.append(f'| {category} | {brand_name} | - | - | {store_info_cell} |')
    
    return '\n'.join(lines)


def generate_markdown_document(results: List[Dict], vr_links: Dict, brand_category_map: Dict[str, str], output_file: str, summary: bool = False, output_dir: str = None):
    """生成 Markdown 文档"""
    lines = []
    
    if output_dir is None:
        output_dir = os.path.dirname(output_file) if output_file else None
    
    if summary:
        # 精简版
        lines.append('# 店铺商品分析报告（精简版）\n')
        lines.append('本报告为精简版本，包含两部分内容：')
        lines.append('1. **商品分析**: 每个类别选择一个店铺，每个店铺选择5个点位，每个点位选择3个推荐商品和2个不推荐商品')
        lines.append('2. **店铺分析**: 每个类别选择一个店铺，每个店铺选择5个点位\n')
        
        # 第一部分：商品分析（精简版）
        lines.append(generate_summary_product_section(results, vr_links, brand_category_map, output_dir=output_dir))
        
        # 第二部分：店铺分析（精简版）
        lines.append(generate_summary_store_section(results, brand_category_map, output_dir=output_dir))
    else:
        # 完整版
        lines.append('# 店铺商品分析报告\n')
        lines.append('本报告包含两部分内容：')
        lines.append('1. **商品分析**: 按店铺和点位展示所有检测到的商品信息')
        lines.append('2. **店铺分析**: 展示各店铺的整体环境分析信息\n')
        
        # 第一部分：商品分析
        lines.append(generate_product_section(results, vr_links, brand_category_map, output_dir=output_dir))
        
        # 第二部分：店铺分析
        lines.append(generate_store_section(results, brand_category_map, output_dir=output_dir))
    
    # 保存文件
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    print(f'\n[INFO] Markdown 文档已保存: {output_file}')
    return '\n'.join(lines)


def format_product_info_for_excel(product: Dict) -> str:
    """格式化商品信息为 Excel 文本"""
    lines = []
    
    # 商品名称（推荐商品加标记）
    name = product.get('name', '')
    is_recommended = product.get('is_recommended', False)
    if is_recommended:
        name = f"{name} ⭐ (推荐)"
    lines.append(f"名称: {name}")
    
    # 其他信息
    if product.get('type'):
        lines.append(f"类型: {product['type']}")
    
    if product.get('colors'):
        colors = '、'.join(product['colors'])
        lines.append(f"颜色: {colors}")
    
    if product.get('materials'):
        materials = '、'.join(product['materials'])
        lines.append(f"材质: {materials}")
    
    if product.get('location'):
        lines.append(f"位置: {product['location']}")
    
    if product.get('view_direction'):
        direction_map = {'f': '前', 'b': '后', 'l': '左', 'r': '右'}
        direction = direction_map.get(product['view_direction'], product['view_direction'])
        lines.append(f"视图方向: {direction}")
    
    if product.get('position_3d'):
        pos = product['position_3d']
        if pos and isinstance(pos, dict):
            lines.append(f"3D坐标: ({pos.get('x', 0)}, {pos.get('y', 0)}, {pos.get('z', 0)})")
    
    return '\n'.join(lines)


def generate_excel_document(results: List[Dict], vr_links: Dict, brand_category_map: Dict[str, str], output_file: str, summary: bool = False, output_dir: str = None):
    """生成 Excel 文档"""
    if output_dir is None:
        output_dir = os.path.dirname(output_file) if output_file else None
    
    # 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表
    
    # 定义样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 第一部分：商品分析
    ws_product = wb.create_sheet(title="商品分析")
    headers = ['店铺分类', '店铺名称', '店铺点位', '点位图片', '商品信息', 'VR链接']
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_product.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 按分类分组品牌
    category_brands = defaultdict(list)
    for brand_result in results:
        brand_name = brand_result.get('brand', '')
        category = get_brand_category(brand_name, brand_category_map)
        category_brands[category].append(brand_result)
    
    row = 2
    for category in sorted(category_brands.keys()):
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        for brand_result in brand_results:
            brand_name = brand_result.get('brand', '')
            product_results = brand_result.get('product_results', [])
            
            if not product_results:
                continue
            
            if summary:
                # 精简版：每个分类选第一个品牌，每个品牌选5个点位，每个点位选3个推荐和2个不推荐
                if brand_result != brand_results[0]:
                    continue
                selected_points = product_results[:5]
            else:
                selected_points = product_results
            
            for point_result in selected_points:
                panorama_id = point_result.get('panorama_id', '')
                images = point_result.get('images', [])
                products = point_result.get('products', [])
                
                if not products:
                    continue
                
                if summary:
                    # 精简版：每个点位选3个推荐和2个不推荐
                    recommended_products = [p for p in products if p.get('is_recommended', False)]
                    non_recommended_products = [p for p in products if not p.get('is_recommended', False)]
                    selected_recommended = recommended_products[:3]
                    selected_non_recommended = non_recommended_products[:2]
                    products = selected_recommended + selected_non_recommended
                
                # 准备图片（直接插入到 Excel）
                # 先设置单元格文本（如果图片加载失败会显示文本）
                image_links = []
                image_files_to_insert = []
                for img_path in sorted(images):
                    # 先获取复制的图片路径
                    actual_image_path = None
                    if output_dir:
                        # 如果指定了输出目录，图片应该已经被复制到 images 文件夹
                        rel_path = get_image_display_path(img_path, output_dir=output_dir)
                        actual_image_path = os.path.join(output_dir, rel_path)
                    else:
                        # 否则使用原始路径
                        actual_image_path = img_path if os.path.exists(img_path) else None
                    
                    direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                    direction = '未知'
                    for key, value in direction_map.items():
                        if img_path.endswith(key):
                            direction = value
                            break
                    
                    if actual_image_path and os.path.exists(actual_image_path):
                        image_files_to_insert.append((actual_image_path, direction))
                    else:
                        image_links.append(f"{direction}: 图片不存在")
                
                images_cell = '\n'.join(image_links) if image_links else '无图片'
                
                # 准备商品信息
                products_info = []
                vr_links_cell = []
                for product in products:
                    product_text = format_product_info_for_excel(product)
                    products_info.append(product_text)
                    
                    # 获取 VR 链接
                    vr_link = None
                    if product.get('is_recommended', False):
                        key = (brand_name, panorama_id, product.get('name', ''))
                        vr_link = vr_links.get(key)
                    
                    if vr_link:
                        vr_links_cell.append(f"{product.get('name', '')}: {vr_link}")
                    elif product.get('is_recommended', False):
                        vr_links_cell.append(f"{product.get('name', '')}: (未找到)")
                
                products_cell = '\n\n'.join(products_info) if products_info else '无商品'
                vr_links_cell_text = '\n'.join(vr_links_cell) if vr_links_cell else '-'
                
                # 写入数据行
                ws_product.cell(row=row, column=1, value=category).alignment = cell_alignment
                ws_product.cell(row=row, column=2, value=brand_name).alignment = cell_alignment
                ws_product.cell(row=row, column=3, value=panorama_id).alignment = cell_alignment
                ws_product.cell(row=row, column=4, value=images_cell).alignment = cell_alignment
                ws_product.cell(row=row, column=5, value=products_cell).alignment = cell_alignment
                ws_product.cell(row=row, column=6, value=vr_links_cell_text).alignment = cell_alignment
                
                # 插入图片到第4列（点位图片）
                if image_files_to_insert:
                    try:
                        # 设置单元格高度（给图片留空间）
                        ws_product.row_dimensions[row].height = 120
                        ws_product.column_dimensions['D'].width = max(ws_product.column_dimensions['D'].width or 30, 35)
                        
                        # 创建组合图片（将4个方向的图片组合成2x2网格）
                        image_size = 60  # 每个小图的高度（单位：像素，Excel中大约对应60点）
                        start_y = 5  # 起始Y位置（从单元格顶部开始）
                        
                        for idx, (img_file, direction) in enumerate(image_files_to_insert[:4]):  # 最多4张图片
                            try:
                                # 打开图片并调整大小
                                pil_img = PILImage.open(img_file)
                                # 计算缩放比例，保持宽高比，高度为 image_size
                                w, h = pil_img.size
                                ratio = image_size / h
                                new_w = int(w * ratio)
                                new_h = image_size
                                
                                # 如果宽度太大，限制宽度
                                if new_w > 80:
                                    ratio = 80 / w
                                    new_w = 80
                                    new_h = int(h * ratio)
                                
                                pil_img_resized = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                                
                                # 保存临时文件
                                temp_dir = output_dir if output_dir else os.path.dirname(img_file)
                                os.makedirs(temp_dir, exist_ok=True)
                                temp_img_path = os.path.join(temp_dir, f'temp_{row}_{idx}.jpg')
                                pil_img_resized.save(temp_img_path, 'JPEG', quality=85)
                                
                                # 创建 Excel 图片对象
                                img = Image(temp_img_path)
                                
                                # 计算位置（2x2网格布局）
                                col_offset = (idx % 2) * (new_w + 2)  # 每行两个，左右排列
                                row_offset = (idx // 2) * (new_h + 2) + start_y  # 每列两个，上下排列
                                
                                # 设置图片位置和大小
                                img.anchor = f'D{row}'
                                img.width = new_w
                                img.height = new_h
                                img.left = col_offset
                                img.top = row_offset
                                
                                # 添加到工作表
                                ws_product.add_image(img)
                                
                                # 删除临时文件
                                try:
                                    os.remove(temp_img_path)
                                except Exception:
                                    pass
                                
                            except Exception as e:
                                print(f'[WARN] 插入图片失败 {img_file}: {e}')
                    except Exception as e:
                        print(f'[WARN] 处理图片时出错: {e}')
                
                # 添加边框
                for col_idx in range(1, 7):
                    ws_product.cell(row=row, column=col_idx).border = border
                
                row += 1
    
    # 设置列宽
    ws_product.column_dimensions['A'].width = 15  # 店铺分类
    ws_product.column_dimensions['B'].width = 20  # 店铺名称
    ws_product.column_dimensions['C'].width = 15  # 店铺点位
    ws_product.column_dimensions['D'].width = 30  # 点位图片
    ws_product.column_dimensions['E'].width = 50  # 商品信息
    ws_product.column_dimensions['F'].width = 50  # VR链接
    
    # 第二部分：店铺分析
    ws_store = wb.create_sheet(title="店铺分析")
    headers_store = ['店铺分类', '店铺名称', '店铺点位', '点位图片', '店铺信息']
    
    # 写入表头
    for col_idx, header in enumerate(headers_store, start=1):
        cell = ws_store.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for category in sorted(category_brands.keys()):
        brand_results = sorted(category_brands[category], key=lambda x: x.get('brand', ''))
        
        for brand_result in brand_results:
            brand_name = brand_result.get('brand', '')
            store_analysis = brand_result.get('store_analysis', {})
            product_results = brand_result.get('product_results', [])
            
            if not store_analysis:
                continue
            
            if summary:
                # 精简版：每个分类选第一个品牌，每个品牌选5个点位
                if brand_result != brand_results[0]:
                    continue
                selected_points = product_results[:5]
            else:
                selected_points = product_results
            
            # 收集点位图片
            all_point_images = {}
            for point_result in selected_points:
                panorama_id = point_result.get('panorama_id', '')
                images = point_result.get('images', [])
                if images:
                    all_point_images[panorama_id] = images
            
            if all_point_images:
                for i, panorama_id in enumerate(sorted(all_point_images.keys())):
                    images = all_point_images[panorama_id]
                    
                    # 准备图片（直接插入到 Excel）
                    image_links = []
                    image_files_to_insert = []
                    for img_path in sorted(images):
                        # 先获取复制的图片路径
                        actual_image_path = None
                        if output_dir:
                            # 如果指定了输出目录，图片应该已经被复制到 images 文件夹
                            rel_path = get_image_display_path(img_path, output_dir=output_dir)
                            actual_image_path = os.path.join(output_dir, rel_path)
                        else:
                            # 否则使用原始路径
                            actual_image_path = img_path if os.path.exists(img_path) else None
                        
                        direction_map = {'_f.jpg': '前', '_b.jpg': '后', '_l.jpg': '左', '_r.jpg': '右'}
                        direction = '未知'
                        for key, value in direction_map.items():
                            if img_path.endswith(key):
                                direction = value
                                break
                        
                        if actual_image_path and os.path.exists(actual_image_path):
                            image_files_to_insert.append((actual_image_path, direction))
                        else:
                            image_links.append(f"{direction}: 图片不存在")
                    
                    images_cell = '\n'.join(image_links) if image_links else '无图片'
                    
                    # 准备店铺信息（第一行显示完整信息）
                    store_info_lines = []
                    if i == 0:
                        if store_analysis.get('store_category'):
                            store_info_lines.append(f"店铺类别: {store_analysis['store_category']}")
                        if store_analysis.get('price_positioning'):
                            store_info_lines.append(f"价格定位: {store_analysis['price_positioning']}")
                        if store_analysis.get('target_customers'):
                            customers = '、'.join(store_analysis['target_customers'])
                            store_info_lines.append(f"目标客户: {customers}")
                        
                        store_env = store_analysis.get('store_env', {})
                        if store_env:
                            if store_env.get('style'):
                                styles = '、'.join(store_env['style']) if isinstance(store_env['style'], list) else store_analysis['store_env']['style']
                                store_info_lines.append(f"风格: {styles}")
                            if store_env.get('lighting'):
                                store_info_lines.append(f"照明: {store_env['lighting']}")
                            if store_env.get('spatial_layout'):
                                store_info_lines.append(f"空间布局: {store_env['spatial_layout']}")
                            if store_env.get('overall_feeling'):
                                store_info_lines.append(f"整体感觉: {store_env['overall_feeling']}")
                            if store_env.get('display_method'):
                                methods = store_env['display_method']
                                if isinstance(methods, list):
                                    methods_text = '\n'.join([f"  - {m}" for m in methods])
                                else:
                                    methods_text = str(methods)
                                store_info_lines.append(f"陈列方式:\n{methods_text}")
                    
                    store_info_cell = '\n'.join(store_info_lines) if store_info_lines else '-'
                    
                    # 写入数据行
                    ws_store.cell(row=row, column=1, value=category).alignment = cell_alignment
                    ws_store.cell(row=row, column=2, value=brand_name).alignment = cell_alignment
                    ws_store.cell(row=row, column=3, value=panorama_id).alignment = cell_alignment
                    ws_store.cell(row=row, column=4, value=images_cell).alignment = cell_alignment
                    ws_store.cell(row=row, column=5, value=store_info_cell).alignment = cell_alignment
                    
                    # 插入图片到第4列（点位图片）
                    if image_files_to_insert:
                        try:
                            # 设置单元格高度（给图片留空间）
                            ws_store.row_dimensions[row].height = 120
                            ws_store.column_dimensions['D'].width = max(ws_store.column_dimensions['D'].width or 30, 35)
                            
                            # 创建组合图片（将4个方向的图片组合成2x2网格）
                            image_size = 60  # 每个小图的高度
                            start_y = 5  # 起始Y位置
                            
                            for idx, (img_file, direction) in enumerate(image_files_to_insert[:4]):  # 最多4张图片
                                try:
                                    # 打开图片并调整大小
                                    pil_img = PILImage.open(img_file)
                                    # 计算缩放比例，保持宽高比
                                    w, h = pil_img.size
                                    ratio = image_size / h
                                    new_w = int(w * ratio)
                                    new_h = image_size
                                    
                                    # 如果宽度太大，限制宽度
                                    if new_w > 80:
                                        ratio = 80 / w
                                        new_w = 80
                                        new_h = int(h * ratio)
                                    
                                    pil_img_resized = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                                    
                                    # 保存临时文件
                                    temp_dir = output_dir if output_dir else os.path.dirname(img_file)
                                    os.makedirs(temp_dir, exist_ok=True)
                                    temp_img_path = os.path.join(temp_dir, f'temp_store_{row}_{idx}.jpg')
                                    pil_img_resized.save(temp_img_path, 'JPEG', quality=85)
                                    
                                    # 创建 Excel 图片对象
                                    img = Image(temp_img_path)
                                    
                                    # 计算位置（2x2网格布局）
                                    col_offset = (idx % 2) * (new_w + 2)
                                    row_offset = (idx // 2) * (new_h + 2) + start_y
                                    
                                    # 设置图片位置和大小
                                    img.anchor = f'D{row}'
                                    img.width = new_w
                                    img.height = new_h
                                    img.left = col_offset
                                    img.top = row_offset
                                    
                                    # 添加到工作表
                                    ws_store.add_image(img)
                                    
                                    # 删除临时文件
                                    try:
                                        os.remove(temp_img_path)
                                    except Exception:
                                        pass
                                    
                                except Exception as e:
                                    print(f'[WARN] 插入图片失败 {img_file}: {e}')
                        except Exception as e:
                            print(f'[WARN] 处理图片时出错: {e}')
                    
                    # 添加边框
                    for col_idx in range(1, 6):
                        ws_store.cell(row=row, column=col_idx).border = border
                    
                    row += 1
    
    # 设置列宽
    ws_store.column_dimensions['A'].width = 15  # 店铺分类
    ws_store.column_dimensions['B'].width = 20  # 店铺名称
    ws_store.column_dimensions['C'].width = 15  # 店铺点位
    ws_store.column_dimensions['D'].width = 30  # 点位图片
    ws_store.column_dimensions['E'].width = 60  # 店铺信息
    
    # 保存文件
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    wb.save(output_file)
    print(f'\n[INFO] Excel 文档已保存: {output_file}')


def markdown_to_html(markdown_content: str) -> str:
    """将 Markdown 内容转换为 HTML"""
    lines = markdown_content.split('\n')
    result_lines = []
    in_table = False
    is_table_header = False
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        
        # 检查是否是表格分隔行
        if stripped.startswith('|---'):
            is_table_header = True
            continue
        
        # 处理表格行
        if stripped.startswith('|'):
            if not in_table:
                result_lines.append('<table class="data-table">')
                in_table = True
                is_table_header = True
            
            # 解析表格单元格
            cells = [cell.strip() for cell in stripped.split('|')[1:-1]]
            
            # 转换单元格内容（处理 Markdown 语法）
            processed_cells = []
            for cell in cells:
                # 处理粗体（优先处理，避免嵌套问题）
                cell = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', cell)
                # 处理链接
                cell = re.sub(r'\[(.+?)\]\((.+?)\)', r'<a href="\2" target="_blank">\1</a>', cell)
                # 处理 bullet points (•) - 保持原样，用 CSS 样式控制
                # 保留 <br> 标签用于换行
                processed_cells.append(cell)
            
            # 输出行
            tag = 'th' if is_table_header else 'td'
            result_lines.append('<tr>')
            for cell in processed_cells:
                result_lines.append(f'<{tag}>{cell}</{tag}>')
            result_lines.append('</tr>')
            is_table_header = False
            continue
        
        # 非表格行
        if in_table:
            result_lines.append('</table>')
            in_table = False
        
        # 处理标题
        if stripped.startswith('#### '):
            title = stripped[5:].strip()
            result_lines.append(f'<h4>{title}</h4>')
        elif stripped.startswith('### '):
            title = stripped[4:].strip()
            result_lines.append(f'<h3>{title}</h3>')
        elif stripped.startswith('## '):
            title = stripped[3:].strip()
            result_lines.append(f'<h2>{title}</h2>')
        elif stripped.startswith('# '):
            title = stripped[2:].strip()
            result_lines.append(f'<h1>{title}</h1>')
        elif stripped:
            # 处理普通段落
            content = line
            # 处理粗体
            content = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', content)
            # 处理链接
            content = re.sub(r'\[(.+?)\]\((.+?)\)', r'<a href="\2" target="_blank">\1</a>', content)
            # 处理列表项（数字列表和 bullet 列表）
            if re.match(r'^\s*\d+\.\s+', content):
                content = re.sub(r'^\s*\d+\.\s+(.+)$', r'<li>\1</li>', content)
                if not any('<ul>' in line or '<ol>' in line for line in result_lines[-5:]):
                    result_lines.append('<ol>')
                result_lines.append(content)
            elif content.strip().startswith('-') or content.strip().startswith('•'):
                content = re.sub(r'^[\s•-]+(.+)$', r'<li>\1</li>', content)
                if not any('<ul>' in line for line in result_lines[-5:]):
                    result_lines.append('<ul>')
                result_lines.append(content)
            else:
                # 检查是否需要关闭列表
                if result_lines and ('</ul>' not in result_lines[-1] and '</ol>' not in result_lines[-1]):
                    if any('<ul>' in line for line in result_lines[-3:]):
                        result_lines.append('</ul>')
                    elif any('<ol>' in line for line in result_lines[-3:]):
                        result_lines.append('</ol>')
                result_lines.append(f'<p>{content}</p>')
        else:
            # 空行 - 检查是否需要关闭列表
            if result_lines:
                if any('<ul>' in line for line in result_lines[-5:]) and '</ul>' not in result_lines[-1]:
                    result_lines.append('</ul>')
                elif any('<ol>' in line for line in result_lines[-5:]) and '</ol>' not in result_lines[-1]:
                    result_lines.append('</ol>')
            result_lines.append('')
    
    if in_table:
        result_lines.append('</table>')
    
    return '\n'.join(result_lines)


def generate_html_document(markdown_content: str, output_file: str):
    """生成 HTML 文档"""
    html_content = markdown_to_html(markdown_content)
    
    # HTML 模板（现代化配色方案）
    html_template = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>店铺商品分析报告</title>
    <style>
        * {{
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", "PingFang SC", "Hiragino Sans GB", Arial, sans-serif;
            line-height: 1.8;
            color: #1a202c;
            max-width: 1600px;
            margin: 0 auto;
            padding: 30px 20px;
            background: #ffffff;
            min-height: 100vh;
        }}
        h1, h2, h3, h4 {{
            text-align: left;
        }}
        p {{
            text-align: left;
        }}
        h1 {{
            color: #1a202c;
            border-bottom: 2px solid #e2e8f0;
            padding: 20px 0 15px 0;
            margin: 40px 0 30px 0;
            font-size: 2.5em;
            font-weight: 700;
            background: white;
            padding-left: 20px;
            padding-right: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }}
        h2 {{
            color: #1a202c;
            background: #f7fafc;
            border: 1px solid #e2e8f0;
            padding: 15px 20px;
            margin: 35px 0 20px 0;
            border-radius: 8px;
            font-size: 1.8em;
            font-weight: 600;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        }}
        h3 {{
            color: #1a202c;
            margin: 25px 0 15px 0;
            font-size: 1.5em;
            font-weight: 600;
            padding-left: 15px;
            border-left: 4px solid #e2e8f0;
        }}
        h4 {{
            color: #718096;
            margin: 20px 0 10px 0;
            font-size: 1.2em;
            font-weight: 500;
        }}
        p {{
            margin: 10px 0;
            padding: 0 10px;
            text-align: left;
        }}
        ul, ol {{
            margin: 10px auto;
            padding-left: 30px;
            display: inline-block;
            text-align: left;
        }}
        li {{
            margin: 8px 0;
            line-height: 1.6;
        }}
        table.data-table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            margin: 25px auto;
            background-color: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 4px 16px rgba(0,0,0,0.1);
        }}
        table.data-table th {{
            background: #f7fafc;
            color: #1a202c;
            padding: 16px 14px;
            text-align: left;
            font-weight: 600;
            font-size: 0.95em;
            letter-spacing: 0.5px;
            border-bottom: 2px solid #e2e8f0;
        }}
        table.data-table th:first-child {{
            border-top-left-radius: 12px;
        }}
        table.data-table th:last-child {{
            border-top-right-radius: 12px;
        }}
        table.data-table td {{
            padding: 14px;
            border-bottom: 1px solid #e2e8f0;
            vertical-align: top;
            background-color: #ffffff;
            color: #1a202c;
            font-size: 0.9em;
            line-height: 1.8;
            text-align: left;
        }}
        table.data-table td p {{
            margin: 8px 0;
            padding: 0;
            color: #1a202c;
        }}
        table.data-table td strong {{
            color: #1a202c;
            font-weight: 700;
            background-color: transparent;
        }}
        table.data-table td .product-label {{
            color: #2d3748;
            font-weight: 600;
        }}
        table.data-table tr:last-child td:first-child {{
            border-bottom-left-radius: 12px;
        }}
        table.data-table tr:last-child td:last-child {{
            border-bottom-right-radius: 12px;
        }}
        table.data-table tr:nth-child(even) td {{
            background-color: #f7fafc;
        }}
        table.data-table tr:hover td {{
            background-color: #edf2f7;
            transition: background-color 0.2s ease;
        }}
        table.data-table tr:last-child td {{
            border-bottom: none;
        }}
        a {{
            color: #1a202c;
            text-decoration: underline;
            font-weight: 600;
            transition: color 0.2s ease;
        }}
        a:hover {{
            color: #2d3748;
            text-decoration: underline;
        }}
        strong {{
            color: #1a202c;
            font-weight: 700;
        }}
        table.data-table td a {{
            color: #1a202c;
            background-color: transparent;
            padding: 0;
            text-decoration: underline;
        }}
        table.data-table td a:hover {{
            color: #2d3748;
            background-color: transparent;
            padding: 0;
        }}
        .category-header {{
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            padding: 20px;
            margin: 25px 0;
            border-radius: 10px;
            box-shadow: 0 4px 12px rgba(245, 87, 108, 0.3);
        }}
    </style>
</head>
<body>
{content}
</body>
</html>"""
    
    full_html = html_template.format(content=html_content)
    
    # 保存 HTML 文件
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(full_html)
    
    print(f'[INFO] HTML 文档已保存: {output_file}')


# ==================== 主函数 ====================
def main():
    parser = argparse.ArgumentParser(
        description='整理分析结果，生成商品分析和店铺分析文档'
    )
    parser.add_argument(
        '--analysis_dir',
        type=str,
        default=DEFAULT_ANALYSIS_DIR,
        help='分析结果目录路径'
    )
    parser.add_argument(
        '--recommended_file',
        type=str,
        default=DEFAULT_RECOMMENDED_FILE,
        help='推荐商品 JSON 文件路径'
    )
    parser.add_argument(
        '--output_dir',
        type=str,
        default=DEFAULT_OUTPUT_DIR,
        help='输出文档目录'
    )
    parser.add_argument(
        '--format',
        type=str,
        choices=['markdown', 'html', 'excel', 'both'],
        default='excel',
        help='输出格式：markdown、html、excel 或 both（默认 excel）'
    )
    parser.add_argument(
        '--summary',
        action='store_true',
        help='同时生成精简版文档（每个类别选一个店铺，每个店铺选5个点）'
    )
    
    args = parser.parse_args()
    
    print('[INFO] 开始整理分析结果...')
    print(f'[INFO] 分析结果目录: {args.analysis_dir}')
    print(f'[INFO] 推荐商品文件: {args.recommended_file}')
    print(f'[INFO] 输出目录: {args.output_dir}')
    print(f'[INFO] 输出格式: {args.format}\n')
    
    # 加载数据
    print('[INFO] 加载数据...')
    brand_category_map = load_brand_categories()
    vr_links = load_recommended_products(args.recommended_file)
    results = load_analysis_results(args.analysis_dir)
    
    if not results:
        print('[ERROR] 未找到任何分析结果文件')
        return
    
    print(f'[INFO] 共加载 {len(results)} 个品牌的分析结果\n')
    
    # 根据 format 参数生成文档
    if args.format in ['excel', 'both']:
        # 生成 Excel 文档
        excel_output_file = os.path.join(args.output_dir, 'analysis_report.xlsx')
        generate_excel_document(results, vr_links, brand_category_map, excel_output_file, summary=False, output_dir=args.output_dir)
        
        # 如果指定了 --summary，同时生成精简版 Excel
        if args.summary:
            print('\n[INFO] 开始生成精简版 Excel 文档...')
            summary_excel_output_file = os.path.join(args.output_dir, 'analysis_report_summary.xlsx')
            generate_excel_document(results, vr_links, brand_category_map, summary_excel_output_file, summary=True, output_dir=args.output_dir)
    
    if args.format in ['markdown', 'html', 'both']:
        # 生成完整版文档
        md_output_file = os.path.join(args.output_dir, 'analysis_report.md')
        html_output_file = os.path.join(args.output_dir, 'analysis_report.html')
        
        markdown_content = None
        
        # 根据 format 参数生成完整版文档
        if args.format in ['markdown', 'both']:
            markdown_content = generate_markdown_document(results, vr_links, brand_category_map, md_output_file, summary=False, output_dir=args.output_dir)
        
        if args.format in ['html', 'both']:
            if markdown_content is None:
                # 如果没有生成 Markdown，需要先生成内容（但不保存文件）
                # 直接调用内部逻辑生成内容
                lines = []
                lines.append('# 店铺商品分析报告\n')
                lines.append('本报告包含两部分内容：')
                lines.append('1. **商品分析**: 按店铺和点位展示所有检测到的商品信息')
                lines.append('2. **店铺分析**: 展示各店铺的整体环境分析信息\n')
                lines.append(generate_product_section(results, vr_links, brand_category_map, output_dir=args.output_dir))
                lines.append(generate_store_section(results, brand_category_map, output_dir=args.output_dir))
                markdown_content = '\n'.join(lines)
            generate_html_document(markdown_content, html_output_file)
        
        print('\n[INFO] 完整版文档生成完成！')
        if args.format in ['markdown', 'both']:
            print(f'  - Markdown: {md_output_file}')
        if args.format in ['html', 'both']:
            print(f'  - HTML: {html_output_file}')
        
        # 如果指定了 --summary，同时生成精简版
        if args.summary:
            print('\n[INFO] 开始生成精简版文档...')
            summary_md_output_file = os.path.join(args.output_dir, 'analysis_report_summary.md')
            summary_html_output_file = os.path.join(args.output_dir, 'analysis_report_summary.html')
            
            summary_markdown_content = None
            
            # 生成精简版 Markdown
            if args.format in ['markdown', 'both']:
                summary_markdown_content = generate_markdown_document(
                    results, vr_links, brand_category_map, summary_md_output_file, summary=True, output_dir=args.output_dir
                )
            
            # 生成精简版 HTML
            if args.format in ['html', 'both']:
                if summary_markdown_content is None:
                    # 如果没有生成 Markdown，需要先生成内容
                    lines = []
                    lines.append('# 店铺商品分析报告（精简版）\n')
                    lines.append('本报告为精简版本，包含两部分内容：')
                    lines.append('1. **商品分析**: 每个类别选择一个店铺，每个店铺选择5个点位，每个点位选择3个推荐商品和2个不推荐商品')
                    lines.append('2. **店铺分析**: 每个类别选择一个店铺，每个店铺选择5个点位\n')
                    lines.append(generate_summary_product_section(results, vr_links, brand_category_map, output_dir=args.output_dir))
                    lines.append(generate_summary_store_section(results, brand_category_map, output_dir=args.output_dir))
                    summary_markdown_content = '\n'.join(lines)
                generate_html_document(summary_markdown_content, summary_html_output_file)
            
            print('\n[INFO] 精简版文档生成完成！')
            if args.format in ['markdown', 'both']:
                print(f'  - Markdown: {summary_md_output_file}')
            if args.format in ['html', 'both']:
                print(f'  - HTML: {summary_html_output_file}')


if __name__ == '__main__':
    main()

