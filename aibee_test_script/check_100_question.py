from datetime import datetime
import sys
import os
import json
import redis

import requests
from datetime import timedelta
from time import sleep, perf_counter
from openpyxl import Workbook, load_workbook
from question import place_questions

# 添加项目根目录到 Python 路径
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))

HISTORY_REDIS_TTL = 60 * 10

class RedisUserContext:
    def __init__(self, place_id, user_id, **kwargs):
        # 创建 Redis 连接
        self._redis = redis_client
        self._key = f"agent_user_context:{place_id}:{user_id}"
        self._ttl = seconds_until_midnight()
        # 存储 session_id 作为 Redis 哈希键
        self.place_id = place_id
        self.user_id = user_id

        for k, v in kwargs.items():
            self.__setattr__(k, v)

    def __setattr__(self, name, value):
        """设置属性时存储到 Redis"""
        # 内部属性直接设置
        if name.startswith('_'):
            super().__setattr__(name, value)
        # 用户属性存储到 Redis
        else:
            try:
                # 序列化值并存储到 Redis 哈希
                serialized = json.dumps(value)
                self._redis.hset(self._key, name, serialized)
                self._redis.expire(self._key, self._ttl)
            except (TypeError, redis.RedisError) as e:
                raise AttributeError(f"Failed to set attribute '{name}': {str(e)}")

    def __getattr__(self, name):
        """获取属性时从 Redis 读取"""
        # 内部属性直接获取
        if name.startswith('_'):
            return super().__getattribute__(name)
        # 从 Redis 哈希读取用户属性
        try:
            serialized = self._redis.hget(self._key, name)
            if serialized is None:
                return None
            return json.loads(serialized)
        except (json.JSONDecodeError, redis.RedisError) as e:
            raise AttributeError(f"Failed to get attribute '{name}': {str(e)}")

    def __delattr__(self, name):
        """删除属性时从 Redis 移除"""
        if name.startswith('_'):
            super().__delattr__(name)
        else:
            if not self._redis.hexists(self._key, name):
                raise AttributeError(f"Attribute '{name}' not found")
            self._redis.hdel(self._key, name)

    def clear(self):
        """清除当前 session_id 的所有属性"""
        self._redis.delete(self._key)

    def get_all(self):
        """获取所有属性字典"""
        try:
            data = self._redis.hgetall(self._key)
            return {k: json.loads(v) for k, v in data.items()}
        except (json.JSONDecodeError, redis.RedisError):
            return {}

def seconds_until_midnight():
    """
    计算从现在到今天凌晨（0点）的秒数。

    Returns:
        float: 从现在到今天凌晨的秒数。
    """
    # 获取当前时间
    now = datetime.now()

    # 获取今天凌晨的时间（即今天0点）
    midnight = datetime(now.year, now.month, now.day) + timedelta(1)

    # 计算从现在到今天凌晨的秒数
    seconds_until_midnight = int((midnight - now).total_seconds())

    return seconds_until_midnight

redis_setting = {
    "REDIS_HOST": "r-bp1glz0v9jhoxl4bqf.redis.rds.aliyuncs.com",
    "REDIS_PORT": 6379,
    "REDIS_PASSWORD": "MpWJZ2roCJ95kyuV",
    "REDIS_DB": 1,
}

redis_client = redis.StrictRedis(host=redis_setting["REDIS_HOST"], port=int(redis_setting["REDIS_PORT"]),
                                 password=redis_setting["REDIS_PASSWORD"], db=int(redis_setting["REDIS_DB"]))


def del_current_intention(place_id, user_id):
    key = f"agent_current_intention:{place_id}_{user_id}"
    redis_client.delete(key)

intentions = [
    ""
]

user_id = "G2GJV8RPoUh35wLd"

filename = "check_100_question.csv"
fieldnames = ["text", "old_intention", "result"]

url_env = {
    'master': "https://screen.aibee.cn/mall_ai_model/question",  # 正式
    'dev': "https://screen.aibee.cn/ai_model/question", # 测试
    'demo': "https://screen.aibee.cn/demo_ai_model/question", # demo
    'local': "http://127.0.0.1:8889/mall_ai_model/question", # 本地
}

def check_100_question(place_id, question, env):
    questions = place_questions.get(place_id, {}).get(question, [])
    t = datetime.now()
    final_result = []
    global_info = RedisUserContext(place_id=place_id, user_id=user_id)
    # 准备按行写入的 Excel 文件
    excel_filename = f"check_100_question_{place_id}_{t.strftime('%Y%m%d%H')}.xlsx"
    if not os.path.exists(excel_filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # 扩展表头，加入计时字段
        ws.append(fieldnames + ["first_token_s", "total_token_s"])
        wb.save(excel_filename)
    total_iters = len(questions) * max(1, len(intentions))
    done_iters = 0

    def print_progress(done: int, total: int):
        total = max(total, 1)
        width = 30
        pct = done / total
        filled = int(width * pct)
        bar = "#" * filled + "." * (width - filled)
        print(f"\r进度 [{bar}] {done}/{total} ({pct*100:.1f}%)", end="", flush=True)
    print_progress(done_iters, total_iters)
    for s in questions:
        for old_intention in intentions:
            sleep(5)
            url = url_env.get(env, "https://screen.aibee.cn/mall_ai_model/question")

            headers = {
                "Content-Type": "application/json",
                "token": user_id,
                "session_id": user_id,
            }

            dataDict = {
                "user_input": s,
                "place_id": place_id
            }
            if old_intention == "":
                del_current_intention(place_id, user_id)
            else:
                redis_client.set(f"agent_current_intention:{place_id}_{user_id}", old_intention, ex=HISTORY_REDIS_TTL)

            data = json.dumps(dataDict)
            start_time = perf_counter()
            first_token_time = None
            response = requests.post(url, data=data, headers=headers, stream=True)
            del_current_intention(place_id, user_id)
            global_info.food_history_user = []
            global_info.food_history_assistant = []
            global_info.historyStoreIDList = []
            global_info.historyStoreType = None
            global_info.exclude_store_list = []
            global_info.old_rewritten_question = None

            res = bytearray()
            for chunk in response.iter_content(chunk_size=None):
                if not chunk:
                    continue
                if first_token_time is None:
                    first_token_time = perf_counter()
                res.extend(chunk)
            result = res.decode('utf-8')
            end_time = perf_counter()

            temp_str = result.split('event:finish')[-1]
            data_str = temp_str.split('data:')[-1]
            sendResult = ""
            try:
                data_list = json.loads(data_str)
                for data in data_list:
                    msg = data.get("msg", None)
                    component = data.get("component", None)
                    intention = data.get("intention", None)
                    if msg:
                        sendResult += msg
                    elif component:
                        sendResult += json.dumps(data, ensure_ascii=False)
                    elif intention:
                        sendResult += json.dumps(data, ensure_ascii=False)
            except json.JSONDecodeError as e:
                print("JSON 解析失败:", e)
            # 改为秒，保留两位小数
            first_token_s = round((first_token_time - start_time), 2) if first_token_time else None
            total_token_s = round((end_time - start_time), 2)
            resDict = {
                "text": s,
                "old_intention": old_intention,
                "result": sendResult,
                "first_token_s": first_token_s,
                "total_token_s": total_token_s,
            }
            print(resDict)
            final_result.append(resDict)
            # 逐行写入 Excel 并即时保存
            try:
                wb = load_workbook(excel_filename)
                ws = wb.active
                # 与表头顺序对应写入
                ws.append([resDict["text"], resDict["old_intention"], resDict["result"], resDict["first_token_s"], resDict["total_token_s"]])
                wb.save(excel_filename)
            except Exception as e:
                print(f"写入 Excel 失败: {e}")
            done_iters += 1
            print_progress(done_iters, total_iters)
            sleep(1)
    print()  # 换行，结束进度条
    print(f"数据已逐行写入文件{excel_filename}")
    print(f"开始时间：{t}")
    print(f"结束时间：{datetime.now()}")
    print(f"耗时：{datetime.now() - t}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Run check_100_question with optional place_id override")
    parser.add_argument("--place_id", type=int, default=801, help="Place ID to override the default in script")
    parser.add_argument("--question", type=str, default='all', help="Input question = all / intention_name")
    parser.add_argument("--env", type=str, default='master', help="Input env = master / dev / demo / local")
    args = parser.parse_args()

    # uv run .\scripts\integration_test\check_100_question.py --question store_introduction --env master
    check_100_question(args.place_id, args.question, args.env)
